from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, field_validator
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import hashlib
import json
import jwt
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
import base64
from enum import Enum

# Windows AD/LDAP Authentication Support
try:
    from ldap3 import Server, Connection, ALL, NTLM, SUBTREE
    from ldap3.core.exceptions import LDAPException, LDAPBindError
    LDAP_AVAILABLE = True
except ImportError:
    LDAP_AVAILABLE = False
    logging.warning("ldap3 not installed - AD authentication disabled")

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'easymoney-secure-key-2024-offline-desktop')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 8

# Master password for app unlock (stored hashed)
MASTER_PASSWORD_HASH_KEY = "master_password_hash"

app = FastAPI(title="EasyMoneyLoans Desktop API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# ==================== ENUMS ====================
class UserRole(str, Enum):
    EMPLOYEE = "employee"
    MANAGER = "manager"
    ADMIN = "admin"

class LoanStatus(str, Enum):
    OPEN = "open"
    PAID = "paid"
    ARCHIVED = "archived"

class RepaymentPlan(int, Enum):
    MONTHLY = 1
    FORTNIGHTLY = 2
    WEEKLY = 4

class AuthMethod(str, Enum):
    LOCAL = "local"
    ACTIVE_DIRECTORY = "active_directory"

# ==================== MODELS ====================
class UserCreate(BaseModel):
    username: str
    password: str
    full_name: str
    role: UserRole
    branch: str

class UserLogin(BaseModel):
    username: str
    password: str
    auth_method: Optional[AuthMethod] = AuthMethod.LOCAL

class ADConfigUpdate(BaseModel):
    enabled: bool = False
    server_url: Optional[str] = None  # e.g., ldap://ad.company.com:389
    domain: Optional[str] = None  # e.g., COMPANY
    base_dn: Optional[str] = None  # e.g., OU=Users,DC=company,DC=com
    default_role: Optional[UserRole] = UserRole.EMPLOYEE
    default_branch: Optional[str] = "Head Office"

class UserResponse(BaseModel):
    id: str
    username: str
    full_name: str
    role: UserRole
    branch: str
    is_active: bool
    created_at: str

class CustomerCreate(BaseModel):
    client_name: str
    id_number: str
    mandate_id: str
    cell_phone: Optional[str] = None
    
    @field_validator('id_number')
    @classmethod
    def validate_sa_id(cls, v):
        if len(v) != 13 or not v.isdigit():
            raise ValueError('SA ID must be exactly 13 digits')
        # Luhn algorithm check
        total = 0
        for i, digit in enumerate(v):
            d = int(digit)
            if i % 2 == 1:
                d *= 2
                if d > 9:
                    d -= 9
            total += d
        if total % 10 != 0:
            raise ValueError('Invalid SA ID checksum')
        return v

class CustomerResponse(BaseModel):
    id: str
    client_name: str
    id_number: str
    id_number_masked: str
    mandate_id: str
    created_at: str
    created_by: str
    created_by_name: str
    updated_at: Optional[str] = None
    updated_by: Optional[str] = None
    archived_at: Optional[str] = None
    archived_by: Optional[str] = None

class LoanCreate(BaseModel):
    customer_id: str
    principal_amount: float
    repayment_plan_code: RepaymentPlan
    loan_date: str

class PaymentResponse(BaseModel):
    id: str
    loan_id: str
    installment_number: int
    amount_due: float
    due_date: str
    is_paid: bool
    paid_at: Optional[str] = None
    paid_by: Optional[str] = None
    paid_by_name: Optional[str] = None
    created_at: str

class LoanResponse(BaseModel):
    id: str
    customer_id: str
    customer_name: str
    customer_id_number: str
    customer_id_number_masked: str
    mandate_id: str
    loan_date: str
    principal_amount: float
    interest_rate: float
    service_fee: float
    total_repayable: float
    repayment_plan_code: int
    installment_amount: float
    outstanding_balance: float
    status: str
    fields_locked: bool
    created_at: str
    created_by: str
    created_by_name: str
    updated_at: Optional[str] = None
    updated_by: Optional[str] = None
    archived_at: Optional[str] = None
    archived_by: Optional[str] = None
    payments: List[PaymentResponse]
    fraud_flags: List[str]

class MarkPaymentRequest(BaseModel):
    loan_id: str
    installment_number: int

class AuditLogResponse(BaseModel):
    id: str
    entity_type: str
    entity_id: str
    action: str
    before_json: Optional[Dict] = None
    after_json: Optional[Dict] = None
    actor_user_id: str
    actor_name: str
    reason: Optional[str] = None
    created_at: str
    integrity_hash: str

class FieldOverrideRequest(BaseModel):
    loan_id: str
    field_name: str
    new_value: Any
    reason: str

class MasterPasswordSetup(BaseModel):
    password: str

class MasterPasswordVerify(BaseModel):
    password: str

class SettingsUpdate(BaseModel):
    export_folder_path: Optional[str] = None
    branch_name: Optional[str] = None

class ExportRequest(BaseModel):
    export_type: str  # 'customers', 'loans', 'payments', 'all'
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    save_to_path: bool = False  # If true, save to configured export folder

class ArchiveRequest(BaseModel):
    entity_type: str  # 'customer', 'loan'
    entity_id: str
    reason: str

# ==================== HELPERS ====================
def mask_id_number(id_number: str) -> str:
    """Mask SA ID: show first 4 and last 3 digits"""
    if len(id_number) != 13:
        return "***********"
    return f"{id_number[:4]}******{id_number[-3:]}"

def calculate_loan(principal: float, plan_code: int) -> dict:
    """Calculate loan details with fixed 40% interest and R12 service fee"""
    interest_rate = 0.40
    service_fee = 12.0
    total_repayable = (principal * (1 + interest_rate)) + service_fee
    installment_amount = total_repayable / plan_code
    return {
        "interest_rate": interest_rate,
        "service_fee": service_fee,
        "total_repayable": round(total_repayable, 2),
        "installment_amount": round(installment_amount, 2)
    }

def generate_payment_schedule(loan_date: str, total: float, plan_code: int) -> list:
    """Generate payment schedule based on plan code"""
    from datetime import datetime
    base_date = datetime.fromisoformat(loan_date.replace('Z', '+00:00'))
    installment = round(total / plan_code, 2)
    payments = []
    
    if plan_code == 1:  # Monthly
        interval_days = 30
    elif plan_code == 2:  # Fortnightly
        interval_days = 14
    else:  # Weekly
        interval_days = 7
    
    for i in range(plan_code):
        due_date = base_date + timedelta(days=interval_days * (i + 1))
        payments.append({
            "installment_number": i + 1,
            "amount_due": installment,
            "due_date": due_date.isoformat(),
            "is_paid": False,
            "paid_at": None,
            "paid_by": None
        })
    return payments

def compute_integrity_hash(data: dict, previous_hash: str = "") -> str:
    """Compute SHA-256 hash for audit log chain"""
    content = json.dumps(data, sort_keys=True, default=str) + previous_hash
    return hashlib.sha256(content.encode()).hexdigest()

async def get_previous_audit_hash() -> str:
    """Get the hash of the most recent audit log entry"""
    last_entry = await db.audit_logs.find_one(
        {},
        sort=[("created_at", -1)],
        projection={"integrity_hash": 1}
    )
    return last_entry["integrity_hash"] if last_entry else ""

async def create_audit_log(entity_type: str, entity_id: str, action: str, 
                           actor_id: str, actor_name: str,
                           before: dict = None, after: dict = None, reason: str = None):
    """Create immutable audit log entry with hash chain"""
    previous_hash = await get_previous_audit_hash()
    
    log_data = {
        "entity_type": entity_type,
        "entity_id": entity_id,
        "action": action,
        "before_json": before,
        "after_json": after,
        "actor_user_id": actor_id,
        "actor_name": actor_name,
        "reason": reason,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    integrity_hash = compute_integrity_hash(log_data, previous_hash)
    log_data["id"] = str(uuid.uuid4())
    log_data["integrity_hash"] = integrity_hash
    
    await db.audit_logs.insert_one(log_data)
    return log_data

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Validate JWT and return current user"""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user or not user.get("is_active"):
            raise HTTPException(status_code=401, detail="User not found or inactive")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(*roles):
    """Dependency to check user role"""
    async def role_checker(user: dict = Depends(get_current_user)):
        if user["role"] not in [r.value for r in roles]:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

# ==================== MASTER PASSWORD ====================
@api_router.get("/master-password/status")
async def check_master_password_status():
    """Check if master password is set"""
    settings = await db.settings.find_one({"key": MASTER_PASSWORD_HASH_KEY})
    return {"is_set": settings is not None}

@api_router.post("/master-password/setup")
async def setup_master_password(data: MasterPasswordSetup):
    """Set up master password (only if not already set)"""
    existing = await db.settings.find_one({"key": MASTER_PASSWORD_HASH_KEY})
    if existing:
        raise HTTPException(status_code=400, detail="Master password already set")
    
    hashed = bcrypt.hashpw(data.password.encode(), bcrypt.gensalt()).decode()
    await db.settings.insert_one({
        "key": MASTER_PASSWORD_HASH_KEY,
        "value": hashed,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Create default admin user
    admin_password = bcrypt.hashpw("admin123".encode(), bcrypt.gensalt()).decode()
    admin_user = {
        "id": str(uuid.uuid4()),
        "username": "admin",
        "password_hash": admin_password,
        "full_name": "System Administrator",
        "role": UserRole.ADMIN.value,
        "branch": "Head Office",
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_user)
    
    return {"message": "Master password set successfully", "default_admin": {"username": "admin", "password": "admin123"}}

@api_router.post("/master-password/verify")
async def verify_master_password(data: MasterPasswordVerify):
    """Verify master password to unlock app"""
    settings = await db.settings.find_one({"key": MASTER_PASSWORD_HASH_KEY})
    if not settings:
        raise HTTPException(status_code=400, detail="Master password not set")
    
    if not bcrypt.checkpw(data.password.encode(), settings["value"].encode()):
        raise HTTPException(status_code=401, detail="Invalid master password")
    
    return {"verified": True}

# ==================== ACTIVE DIRECTORY AUTHENTICATION ====================
async def authenticate_with_ad(username: str, password: str, ad_config: dict) -> Optional[dict]:
    """
    Authenticate user against Windows Active Directory.
    Returns user info dict if successful, None if failed.
    """
    if not LDAP_AVAILABLE:
        logging.error("LDAP library not available")
        return None
    
    if not ad_config or not ad_config.get("enabled"):
        return None
    
    server_url = ad_config.get("server_url")
    domain = ad_config.get("domain")
    base_dn = ad_config.get("base_dn")
    
    if not all([server_url, domain]):
        logging.error("AD configuration incomplete")
        return None
    
    try:
        # Build the user DN for authentication
        # Support formats: username, DOMAIN\username, username@domain.com
        if '\\' not in username and '@' not in username:
            user_dn = f"{domain}\\{username}"
        else:
            user_dn = username
        
        server = Server(server_url, get_info=ALL)
        
        # Extract clean username for search
        clean_username = username.replace('\\', '/').split('/')[-1].split('@')[0]
        
        # Try NTLM authentication (most common for Windows AD)
        with Connection(server, user=user_dn, password=password, authentication=NTLM, raise_exceptions=True) as conn:
            if conn.bound:
                # Successfully authenticated, now get user info
                search_filter = f"(sAMAccountName={clean_username})"
                
                if base_dn:
                    conn.search(
                        search_base=base_dn,
                        search_filter=search_filter,
                        search_scope=SUBTREE,
                        attributes=['sAMAccountName', 'displayName', 'mail', 'givenName', 'sn', 'memberOf']
                    )
                    
                    if conn.entries:
                        entry = conn.entries[0]
                        return {
                            'username': str(entry.sAMAccountName) if entry.sAMAccountName else username,
                            'full_name': str(entry.displayName) if entry.displayName else username,
                            'email': str(entry.mail) if entry.mail else None,
                            'groups': [str(g) for g in entry.memberOf] if entry.memberOf else []
                        }
                
                # Fallback: Return basic info if search fails
                return {
                    'username': clean_username,
                    'full_name': clean_username,
                    'email': None,
                    'groups': []
                }
        
        return None
    
    except LDAPBindError as e:
        logging.warning(f"AD bind failed for user {username}: {e}")
        return None
    except LDAPException as e:
        logging.error(f"LDAP error during AD authentication: {e}")
        return None
    except Exception as e:
        logging.error(f"Unexpected error during AD authentication: {e}")
        return None

# ==================== AUTH ====================
@api_router.post("/auth/login")
async def login(data: UserLogin):
    """User login - supports local and Active Directory authentication"""
    
    # Get AD configuration
    ad_config = await db.settings.find_one({"key": "ad_config"})
    ad_settings = ad_config.get("value", {}) if ad_config else {}
    
    user = None
    ad_authenticated = False
    
    # Try AD authentication if enabled and requested
    if data.auth_method == AuthMethod.ACTIVE_DIRECTORY or (ad_settings.get("enabled") and data.auth_method != AuthMethod.LOCAL):
        ad_user_info = await authenticate_with_ad(data.username, data.password, ad_settings)
        
        if ad_user_info:
            ad_authenticated = True
            # Check if user exists in local database
            user = await db.users.find_one({"username": ad_user_info['username']}, {"_id": 0})
            
            if not user:
                # Auto-create user from AD
                user = {
                    "id": str(uuid.uuid4()),
                    "username": ad_user_info['username'],
                    "password_hash": "",  # No local password for AD users
                    "full_name": ad_user_info['full_name'],
                    "role": ad_settings.get("default_role", UserRole.EMPLOYEE.value),
                    "branch": ad_settings.get("default_branch", "Head Office"),
                    "is_active": True,
                    "auth_method": AuthMethod.ACTIVE_DIRECTORY.value,
                    "ad_groups": ad_user_info.get('groups', []),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.users.insert_one(user)
                await create_audit_log("user", user["id"], "ad_auto_create", user["id"], user["full_name"], after={"username": user["username"]})
            else:
                # Update AD info
                await db.users.update_one(
                    {"id": user["id"]},
                    {"$set": {
                        "full_name": ad_user_info['full_name'],
                        "ad_groups": ad_user_info.get('groups', []),
                        "last_ad_sync": datetime.now(timezone.utc).isoformat()
                    }}
                )
    
    # Fall back to local authentication if AD not used or failed
    if not ad_authenticated:
        user = await db.users.find_one({"username": data.username}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        if not user.get("password_hash"):
            raise HTTPException(status_code=401, detail="This user can only authenticate via Active Directory")
        
        if not bcrypt.checkpw(data.password.encode(), user["password_hash"].encode()):
            raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user.get("is_active"):
        raise HTTPException(status_code=401, detail="Account disabled")
    
    token = jwt.encode({
        "sub": user["id"],
        "username": user["username"],
        "role": user["role"],
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }, JWT_SECRET, algorithm=JWT_ALGORITHM)
    
    await create_audit_log("user", user["id"], "login", user["id"], user["full_name"], 
                           after={"auth_method": "ad" if ad_authenticated else "local"})
    
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "full_name": user["full_name"],
            "role": user["role"],
            "branch": user["branch"]
        },
        "auth_method": "active_directory" if ad_authenticated else "local"
    }

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    """Get current user info"""
    return {
        "id": user["id"],
        "username": user["username"],
        "full_name": user["full_name"],
        "role": user["role"],
        "branch": user["branch"]
    }

# ==================== USERS (Admin only) ====================
@api_router.get("/users", response_model=List[UserResponse])
async def list_users(user: dict = Depends(require_role(UserRole.ADMIN))):
    """List all users"""
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.post("/users", response_model=UserResponse)
async def create_user(data: UserCreate, user: dict = Depends(require_role(UserRole.ADMIN))):
    """Create new user"""
    existing = await db.users.find_one({"username": data.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    password_hash = bcrypt.hashpw(data.password.encode(), bcrypt.gensalt()).decode()
    new_user = {
        "id": str(uuid.uuid4()),
        "username": data.username,
        "password_hash": password_hash,
        "full_name": data.full_name,
        "role": data.role.value,
        "branch": data.branch,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(new_user)
    await create_audit_log("user", new_user["id"], "create", user["id"], user["full_name"], after={"username": data.username, "role": data.role.value})
    
    return {k: v for k, v in new_user.items() if k != "password_hash"}

@api_router.put("/users/{user_id}/toggle-active")
async def toggle_user_active(user_id: str, user: dict = Depends(require_role(UserRole.ADMIN))):
    """Toggle user active status"""
    target = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    
    new_status = not target.get("is_active", True)
    await db.users.update_one({"id": user_id}, {"$set": {"is_active": new_status}})
    await create_audit_log("user", user_id, "toggle_active", user["id"], user["full_name"], 
                           before={"is_active": target.get("is_active")}, after={"is_active": new_status})
    
    return {"is_active": new_status}

# ==================== CUSTOMERS ====================
@api_router.get("/customers")
async def list_customers(user: dict = Depends(get_current_user)):
    """List all customers"""
    customers = await db.customers.find({"archived_at": None}, {"_id": 0}).to_list(1000)
    can_view_full_id = user["role"] in [UserRole.MANAGER.value, UserRole.ADMIN.value]
    
    result = []
    for c in customers:
        creator = await db.users.find_one({"id": c.get("created_by")}, {"full_name": 1, "_id": 0})
        result.append({
            **c,
            "id_number": c["id_number"] if can_view_full_id else mask_id_number(c["id_number"]),
            "id_number_masked": mask_id_number(c["id_number"]),
            "created_by_name": creator["full_name"] if creator else "Unknown"
        })
    return result

@api_router.post("/customers")
async def create_customer(data: CustomerCreate, user: dict = Depends(get_current_user)):
    """Create new customer"""
    # Check for duplicate
    existing = await db.customers.find_one({
        "client_name": data.client_name,
        "id_number": data.id_number,
        "archived_at": None
    })
    if existing:
        raise HTTPException(status_code=400, detail="Customer with same name and ID already exists")
    
    customer = {
        "id": str(uuid.uuid4()),
        "client_name": data.client_name,
        "id_number": data.id_number,
        "mandate_id": data.mandate_id,
        "cell_phone": data.cell_phone,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "updated_at": None,
        "updated_by": None,
        "archived_at": None,
        "archived_by": None
    }
    
    await db.customers.insert_one(customer)
    await create_audit_log("customer", customer["id"], "create", user["id"], user["full_name"], 
                           after={"client_name": data.client_name, "mandate_id": data.mandate_id})
    
    creator = await db.users.find_one({"id": user["id"]}, {"full_name": 1, "_id": 0})
    can_view_full_id = user["role"] in [UserRole.MANAGER.value, UserRole.ADMIN.value]
    
    # Remove MongoDB _id from customer dict before returning
    customer_response = {k: v for k, v in customer.items() if k != "_id"}
    
    return {
        **customer_response,
        "id_number": customer["id_number"] if can_view_full_id else mask_id_number(customer["id_number"]),
        "id_number_masked": mask_id_number(customer["id_number"]),
        "created_by_name": creator["full_name"] if creator else "Unknown"
    }

@api_router.get("/customers/{customer_id}")
async def get_customer(customer_id: str, user: dict = Depends(get_current_user)):
    """Get customer details"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    creator = await db.users.find_one({"id": customer.get("created_by")}, {"full_name": 1, "_id": 0})
    can_view_full_id = user["role"] in [UserRole.MANAGER.value, UserRole.ADMIN.value]
    
    return {
        **customer,
        "id_number": customer["id_number"] if can_view_full_id else mask_id_number(customer["id_number"]),
        "id_number_masked": mask_id_number(customer["id_number"]),
        "created_by_name": creator["full_name"] if creator else "Unknown"
    }

# ==================== LOANS ====================
@api_router.get("/loans")
async def list_loans(
    loan_status: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """List all loans with fraud detection"""
    query = {"archived_at": None}
    if loan_status:
        query["status"] = loan_status
    
    loans = await db.loans.find(query, {"_id": 0}).to_list(1000)
    can_view_full_id = user["role"] in [UserRole.MANAGER.value, UserRole.ADMIN.value]
    
    # Detect duplicates (same customer with multiple loans)
    customer_loan_count = {}
    for loan in loans:
        cid = loan["customer_id"]
        customer_loan_count[cid] = customer_loan_count.get(cid, 0) + 1
    
    result = []
    for loan in loans:
        customer = await db.customers.find_one({"id": loan["customer_id"]}, {"_id": 0})
        creator = await db.users.find_one({"id": loan.get("created_by")}, {"full_name": 1, "_id": 0})
        payments = await db.payments.find({"loan_id": loan["id"]}, {"_id": 0}).to_list(100)
        
        # Enrich payments with paid_by_name
        enriched_payments = []
        for p in payments:
            paid_by_name = None
            if p.get("paid_by"):
                payer = await db.users.find_one({"id": p["paid_by"]}, {"full_name": 1, "_id": 0})
                paid_by_name = payer["full_name"] if payer else "Unknown"
            enriched_payments.append({**p, "paid_by_name": paid_by_name})
        
        # Fraud flags
        fraud_flags = []
        
        # Quick-close detection: created and fully paid same day
        if loan["status"] == LoanStatus.PAID.value:
            created_date = loan["created_at"][:10]
            last_payment = max([p["paid_at"][:10] for p in payments if p.get("paid_at")] or [None])
            if last_payment and created_date == last_payment:
                fraud_flags.append("QUICK_CLOSE")
        
        # Duplicate customer
        if customer_loan_count.get(loan["customer_id"], 0) > 1:
            fraud_flags.append("DUPLICATE_CUSTOMER")
        
        result.append({
            **loan,
            "customer_name": customer["client_name"] if customer else "Unknown",
            "customer_id_number": customer["id_number"] if customer and can_view_full_id else mask_id_number(customer["id_number"]) if customer else "Unknown",
            "customer_id_number_masked": mask_id_number(customer["id_number"]) if customer else "Unknown",
            "customer_mandate_id": customer.get("mandate_id") if customer else None,
            "customer_sassa_end": customer.get("sassa_end_date") if customer else None,
            "mandate_id": customer["mandate_id"] if customer else "Unknown",
            "created_by_name": creator["full_name"] if creator else "Unknown",
            "payments": enriched_payments,
            "fraud_flags": fraud_flags
        })
    
    return result

@api_router.post("/loans")
async def create_loan(data: LoanCreate, user: dict = Depends(get_current_user)):
    """Create new loan"""
    customer = await db.customers.find_one({"id": data.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    calc = calculate_loan(data.principal_amount, data.repayment_plan_code.value)
    payments = generate_payment_schedule(data.loan_date, calc["total_repayable"], data.repayment_plan_code.value)
    
    loan = {
        "id": str(uuid.uuid4()),
        "customer_id": data.customer_id,
        "loan_date": data.loan_date,
        "principal_amount": data.principal_amount,
        "interest_rate": calc["interest_rate"],
        "service_fee": calc["service_fee"],
        "total_repayable": calc["total_repayable"],
        "repayment_plan_code": data.repayment_plan_code.value,
        "installment_amount": calc["installment_amount"],
        "outstanding_balance": calc["total_repayable"],
        "status": LoanStatus.OPEN.value,
        "fields_locked": True,  # Lock immediately after creation
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "updated_at": None,
        "updated_by": None,
        "archived_at": None,
        "archived_by": None
    }
    
    await db.loans.insert_one(loan)
    
    # Create payment records
    for p in payments:
        payment_doc = {
            "id": str(uuid.uuid4()),
            "loan_id": loan["id"],
            **p,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.payments.insert_one(payment_doc)
    
    await create_audit_log("loan", loan["id"], "create", user["id"], user["full_name"],
                           after={"customer_id": data.customer_id, "principal": data.principal_amount, "plan": data.repayment_plan_code.value})
    
    return {"id": loan["id"], "message": "Loan created successfully"}

@api_router.get("/loans/{loan_id}")
async def get_loan(loan_id: str, user: dict = Depends(get_current_user)):
    """Get loan details"""
    loan = await db.loans.find_one({"id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    customer = await db.customers.find_one({"id": loan["customer_id"]}, {"_id": 0})
    creator = await db.users.find_one({"id": loan.get("created_by")}, {"full_name": 1, "_id": 0})
    payments = await db.payments.find({"loan_id": loan_id}, {"_id": 0}).to_list(100)
    can_view_full_id = user["role"] in [UserRole.MANAGER.value, UserRole.ADMIN.value]
    
    # Enrich payments
    enriched_payments = []
    for p in payments:
        paid_by_name = None
        if p.get("paid_by"):
            payer = await db.users.find_one({"id": p["paid_by"]}, {"full_name": 1, "_id": 0})
            paid_by_name = payer["full_name"] if payer else "Unknown"
        enriched_payments.append({**p, "paid_by_name": paid_by_name})
    
    # Check fraud flags
    fraud_flags = []
    all_loans = await db.loans.find({"customer_id": loan["customer_id"], "archived_at": None}, {"_id": 0}).to_list(100)
    if len(all_loans) > 1:
        fraud_flags.append("DUPLICATE_CUSTOMER")
    
    if loan["status"] == LoanStatus.PAID.value:
        created_date = loan["created_at"][:10]
        last_payment = max([p["paid_at"][:10] for p in payments if p.get("paid_at")] or [None])
        if last_payment and created_date == last_payment:
            fraud_flags.append("QUICK_CLOSE")
    
    return {
        **loan,
        "customer_name": customer["client_name"] if customer else "Unknown",
        "customer_id_number": customer["id_number"] if customer and can_view_full_id else mask_id_number(customer["id_number"]) if customer else "Unknown",
        "customer_id_number_masked": mask_id_number(customer["id_number"]) if customer else "Unknown",
        "mandate_id": customer["mandate_id"] if customer else "Unknown",
        "created_by_name": creator["full_name"] if creator else "Unknown",
        "payments": enriched_payments,
        "fraud_flags": fraud_flags
    }

# ==================== PAYMENTS ====================
@api_router.post("/payments/mark-paid")
async def mark_payment_paid(data: MarkPaymentRequest, user: dict = Depends(get_current_user)):
    """Mark a payment as paid - Employee can mark payments"""
    payment = await db.payments.find_one({
        "loan_id": data.loan_id,
        "installment_number": data.installment_number
    }, {"_id": 0})
    
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    if payment.get("is_paid"):
        raise HTTPException(status_code=400, detail="Payment already marked as paid - cannot be reversed")
    
    # Mark as paid
    now = datetime.now(timezone.utc).isoformat()
    await db.payments.update_one(
        {"id": payment["id"]},
        {"$set": {"is_paid": True, "paid_at": now, "paid_by": user["id"]}}
    )
    
    # Update loan outstanding balance
    loan = await db.loans.find_one({"id": data.loan_id}, {"_id": 0})
    new_balance = round(loan["outstanding_balance"] - payment["amount_due"], 2)
    new_status = LoanStatus.PAID.value if new_balance <= 0 else loan["status"]
    
    await db.loans.update_one(
        {"id": data.loan_id},
        {"$set": {
            "outstanding_balance": max(0, new_balance),
            "status": new_status,
            "updated_at": now,
            "updated_by": user["id"]
        }}
    )
    
    await create_audit_log("payment", payment["id"], "mark_paid", user["id"], user["full_name"],
                           before={"is_paid": False}, after={"is_paid": True, "paid_at": now})
    
    return {"message": "Payment marked as paid", "new_balance": max(0, new_balance), "loan_status": new_status}

# ==================== FIELD OVERRIDE (Manager/Admin) ====================
@api_router.post("/loans/override-field")
async def override_loan_field(data: FieldOverrideRequest, user: dict = Depends(require_role(UserRole.MANAGER, UserRole.ADMIN))):
    """Override locked loan field (Manager/Admin only with reason)"""
    loan = await db.loans.find_one({"id": data.loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    allowed_fields = ["loan_date", "principal_amount", "repayment_plan_code"]
    if data.field_name not in allowed_fields:
        raise HTTPException(status_code=400, detail=f"Field {data.field_name} cannot be overridden")
    
    if not data.reason or len(data.reason) < 10:
        raise HTTPException(status_code=400, detail="Reason must be at least 10 characters")
    
    before_value = loan.get(data.field_name)
    
    # Recalculate if principal or plan changes
    update_fields = {data.field_name: data.new_value, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": user["id"]}
    
    if data.field_name in ["principal_amount", "repayment_plan_code"]:
        principal = data.new_value if data.field_name == "principal_amount" else loan["principal_amount"]
        plan = data.new_value if data.field_name == "repayment_plan_code" else loan["repayment_plan_code"]
        calc = calculate_loan(principal, plan)
        update_fields.update(calc)
        update_fields["outstanding_balance"] = calc["total_repayable"]
    
    await db.loans.update_one({"id": data.loan_id}, {"$set": update_fields})
    
    await create_audit_log("loan", data.loan_id, "field_override", user["id"], user["full_name"],
                           before={data.field_name: before_value}, after={data.field_name: data.new_value}, reason=data.reason)
    
    return {"message": f"Field {data.field_name} updated successfully"}

# ==================== EXPORT (Manager/Admin) ====================
@api_router.post("/export")
async def export_data(data: ExportRequest, user: dict = Depends(require_role(UserRole.MANAGER, UserRole.ADMIN))):
    """Export data to Excel format (returns base64)"""
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if data.export_type in ["customers", "all"]:
        ws = wb.active if data.export_type == "customers" else wb.create_sheet("Customers")
        ws.title = "Customers"
        
        headers = ["ID", "Client Name", "ID Number", "Mandate ID", "Created At", "Created By"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        customers = await db.customers.find({"archived_at": None}, {"_id": 0}).to_list(10000)
        for row, c in enumerate(customers, 2):
            creator = await db.users.find_one({"id": c.get("created_by")}, {"full_name": 1, "_id": 0})
            ws.cell(row=row, column=1, value=c["id"]).border = border
            ws.cell(row=row, column=2, value=c["client_name"]).border = border
            # Format ID as text to prevent scientific notation
            id_cell = ws.cell(row=row, column=3, value=c["id_number"])
            id_cell.number_format = '@'
            id_cell.border = border
            ws.cell(row=row, column=4, value=c["mandate_id"]).border = border
            ws.cell(row=row, column=5, value=c["created_at"]).border = border
            ws.cell(row=row, column=6, value=creator["full_name"] if creator else "Unknown").border = border
        
        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    if data.export_type in ["loans", "all"]:
        ws = wb.active if data.export_type == "loans" else wb.create_sheet("Loans")
        if data.export_type == "loans":
            ws.title = "Loans"
        
        headers = ["Loan ID", "Customer Name", "Customer ID", "Principal", "Total Repayable", 
                   "Outstanding", "Status", "Plan", "Created At", "Created By"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        loans = await db.loans.find({"archived_at": None}, {"_id": 0}).to_list(10000)
        for row, loan in enumerate(loans, 2):
            customer = await db.customers.find_one({"id": loan["customer_id"]}, {"_id": 0})
            creator = await db.users.find_one({"id": loan.get("created_by")}, {"full_name": 1, "_id": 0})
            
            plan_names = {1: "Monthly", 2: "Fortnightly", 4: "Weekly"}
            
            ws.cell(row=row, column=1, value=loan["id"]).border = border
            ws.cell(row=row, column=2, value=customer["client_name"] if customer else "Unknown").border = border
            id_cell = ws.cell(row=row, column=3, value=customer["id_number"] if customer else "Unknown")
            id_cell.number_format = '@'
            id_cell.border = border
            ws.cell(row=row, column=4, value=f"R{loan['principal_amount']:.2f}").border = border
            ws.cell(row=row, column=5, value=f"R{loan['total_repayable']:.2f}").border = border
            ws.cell(row=row, column=6, value=f"R{loan['outstanding_balance']:.2f}").border = border
            ws.cell(row=row, column=7, value=loan["status"].upper()).border = border
            ws.cell(row=row, column=8, value=plan_names.get(loan["repayment_plan_code"], "Unknown")).border = border
            ws.cell(row=row, column=9, value=loan["created_at"]).border = border
            ws.cell(row=row, column=10, value=creator["full_name"] if creator else "Unknown").border = border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    if data.export_type in ["payments", "all"]:
        ws = wb.create_sheet("Payments") if data.export_type == "all" else wb.active
        if data.export_type == "payments":
            ws.title = "Payments"
        
        headers = ["Payment ID", "Loan ID", "Installment #", "Amount Due", "Due Date", 
                   "Is Paid", "Paid At", "Paid By"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        payments = await db.payments.find({}, {"_id": 0}).to_list(50000)
        for row, p in enumerate(payments, 2):
            payer = None
            if p.get("paid_by"):
                payer = await db.users.find_one({"id": p["paid_by"]}, {"full_name": 1, "_id": 0})
            
            ws.cell(row=row, column=1, value=p["id"]).border = border
            ws.cell(row=row, column=2, value=p["loan_id"]).border = border
            ws.cell(row=row, column=3, value=p["installment_number"]).border = border
            ws.cell(row=row, column=4, value=f"R{p['amount_due']:.2f}").border = border
            ws.cell(row=row, column=5, value=p["due_date"]).border = border
            ws.cell(row=row, column=6, value="Yes" if p["is_paid"] else "No").border = border
            ws.cell(row=row, column=7, value=p.get("paid_at", "")).border = border
            ws.cell(row=row, column=8, value=payer["full_name"] if payer else "").border = border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Remove default empty sheet if exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    
    # Generate filename
    branch = user.get("branch", "Unknown").replace(" ", "_")
    filename = f"Loans_{datetime.now().strftime('%Y-%m-%d')}_{branch}.xlsx"
    
    # Check if we should save to configured export folder
    if data.save_to_path:
        settings = await db.settings.find_one({"key": "export_folder_path"})
        if settings and settings.get("value"):
            export_path = settings["value"]
            # Ensure directory exists
            import os
            if os.path.isdir(export_path):
                full_path = os.path.join(export_path, filename)
                wb.save(full_path)
                
                await create_audit_log("export", "system", "export_data", user["id"], user["full_name"],
                                       after={"export_type": data.export_type, "saved_to": full_path})
                
                return {
                    "filename": filename,
                    "saved_to_path": full_path,
                    "message": f"Export saved to {full_path}"
                }
            else:
                raise HTTPException(status_code=400, detail=f"Export folder does not exist: {export_path}")
        else:
            raise HTTPException(status_code=400, detail="Export folder path not configured. Please configure in Admin Settings.")
    
    # Save to bytes for download
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    await create_audit_log("export", "system", "export_data", user["id"], user["full_name"],
                           after={"export_type": data.export_type})
    
    return {
        "filename": filename,
        "data": base64.b64encode(buffer.getvalue()).decode(),
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

# ==================== ARCHIVE (Admin only) ====================
@api_router.post("/archive")
async def archive_entity(data: ArchiveRequest, user: dict = Depends(require_role(UserRole.ADMIN))):
    """Archive customer or loan (Admin only)"""
    if not data.reason or len(data.reason) < 10:
        raise HTTPException(status_code=400, detail="Archive reason must be at least 10 characters")
    
    now = datetime.now(timezone.utc).isoformat()
    
    if data.entity_type == "customer":
        entity = await db.customers.find_one({"id": data.entity_id}, {"_id": 0})
        if not entity:
            raise HTTPException(status_code=404, detail="Customer not found")
        await db.customers.update_one({"id": data.entity_id}, {"$set": {"archived_at": now, "archived_by": user["id"]}})
    elif data.entity_type == "loan":
        entity = await db.loans.find_one({"id": data.entity_id}, {"_id": 0})
        if not entity:
            raise HTTPException(status_code=404, detail="Loan not found")
        await db.loans.update_one({"id": data.entity_id}, {"$set": {"archived_at": now, "archived_by": user["id"]}})
    else:
        raise HTTPException(status_code=400, detail="Invalid entity type")
    
    await create_audit_log(data.entity_type, data.entity_id, "archive", user["id"], user["full_name"], reason=data.reason)
    
    return {"message": f"{data.entity_type.capitalize()} archived successfully"}

# ==================== AUDIT LOGS ====================
@api_router.get("/audit-logs")
async def list_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    actor_id: Optional[str] = None,
    limit: int = 100,
    user: dict = Depends(get_current_user)
):
    """List audit logs (read-only)"""
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    if actor_id:
        query["actor_user_id"] = actor_id
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return logs

@api_router.get("/audit-logs/verify-integrity")
async def verify_audit_integrity(user: dict = Depends(require_role(UserRole.ADMIN))):
    """Verify audit log chain integrity"""
    logs = await db.audit_logs.find({}, {"_id": 0}).sort("created_at", 1).to_list(100000)
    
    if not logs:
        return {"valid": True, "message": "No audit logs to verify"}
    
    previous_hash = ""
    invalid_entries = []
    
    for log in logs:
        stored_hash = log.pop("integrity_hash")
        log_id = log.pop("id")
        computed_hash = compute_integrity_hash(log, previous_hash)
        
        if computed_hash != stored_hash:
            invalid_entries.append({"id": log_id, "expected": computed_hash, "stored": stored_hash})
        
        previous_hash = stored_hash
    
    if invalid_entries:
        return {"valid": False, "message": "Audit log tampering detected!", "invalid_entries": invalid_entries[:5]}
    
    return {"valid": True, "message": f"All {len(logs)} audit log entries verified", "total_entries": len(logs)}

# ==================== SETTINGS ====================
@api_router.get("/settings")
async def get_settings(user: dict = Depends(get_current_user)):
    """Get app settings"""
    settings = await db.settings.find({"key": {"$ne": MASTER_PASSWORD_HASH_KEY}}, {"_id": 0}).to_list(100)
    return {s["key"]: s["value"] for s in settings}

@api_router.put("/settings")
async def update_settings(data: SettingsUpdate, user: dict = Depends(require_role(UserRole.ADMIN))):
    """Update app settings (Admin only)"""
    updates = data.model_dump(exclude_none=True)
    
    for key, value in updates.items():
        await db.settings.update_one(
            {"key": key},
            {"$set": {"key": key, "value": value, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
    
    await create_audit_log("settings", "system", "update", user["id"], user["full_name"], after=updates)
    
    return {"message": "Settings updated successfully"}

# ==================== ACTIVE DIRECTORY CONFIGURATION ====================
@api_router.get("/settings/ad-config")
async def get_ad_config(user: dict = Depends(require_role(UserRole.ADMIN))):
    """Get Active Directory configuration (Admin only)"""
    config = await db.settings.find_one({"key": "ad_config"})
    if config:
        # Mask sensitive data
        value = config.get("value", {})
        return {
            "enabled": value.get("enabled", False),
            "server_url": value.get("server_url", ""),
            "domain": value.get("domain", ""),
            "base_dn": value.get("base_dn", ""),
            "default_role": value.get("default_role", UserRole.EMPLOYEE.value),
            "default_branch": value.get("default_branch", "Head Office"),
            "ldap_available": LDAP_AVAILABLE
        }
    return {
        "enabled": False,
        "server_url": "",
        "domain": "",
        "base_dn": "",
        "default_role": UserRole.EMPLOYEE.value,
        "default_branch": "Head Office",
        "ldap_available": LDAP_AVAILABLE
    }

@api_router.put("/settings/ad-config")
async def update_ad_config(data: ADConfigUpdate, user: dict = Depends(require_role(UserRole.ADMIN))):
    """Update Active Directory configuration (Admin only)"""
    if not LDAP_AVAILABLE:
        raise HTTPException(status_code=400, detail="LDAP library not available - install ldap3 package")
    
    config_value = {
        "enabled": data.enabled,
        "server_url": data.server_url,
        "domain": data.domain,
        "base_dn": data.base_dn,
        "default_role": data.default_role.value if data.default_role else UserRole.EMPLOYEE.value,
        "default_branch": data.default_branch or "Head Office"
    }
    
    await db.settings.update_one(
        {"key": "ad_config"},
        {"$set": {"key": "ad_config", "value": config_value, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    await create_audit_log("settings", "ad_config", "update", user["id"], user["full_name"], 
                           after={"enabled": data.enabled, "server_url": data.server_url})
    
    return {"message": "Active Directory configuration updated successfully"}

@api_router.post("/settings/ad-config/test")
async def test_ad_connection(data: ADConfigUpdate, user: dict = Depends(require_role(UserRole.ADMIN))):
    """Test Active Directory connection (Admin only)"""
    if not LDAP_AVAILABLE:
        raise HTTPException(status_code=400, detail="LDAP library not available")
    
    if not data.server_url or not data.domain:
        raise HTTPException(status_code=400, detail="Server URL and domain are required")
    
    try:
        server = Server(data.server_url, get_info=ALL, connect_timeout=10)
        conn = Connection(server, auto_bind=False)
        conn.open()
        
        server_info = {
            "server_type": str(conn.server.info.vendor_name) if conn.server.info else "Unknown",
            "naming_contexts": list(conn.server.info.naming_contexts) if conn.server.info and conn.server.info.naming_contexts else []
        }
        
        conn.unbind()
        
        return {
            "success": True,
            "message": "Successfully connected to Active Directory server",
            "server_info": server_info
        }
    except Exception as e:
        return {
            "success": False,
            "message": f"Connection failed: {str(e)}",
            "server_info": None
        }

# ==================== DASHBOARD STATS ====================
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    """Get dashboard statistics"""
    total_customers = await db.customers.count_documents({"archived_at": None})
    total_loans = await db.loans.count_documents({"archived_at": None})
    open_loans = await db.loans.count_documents({"status": LoanStatus.OPEN.value, "archived_at": None})
    paid_loans = await db.loans.count_documents({"status": LoanStatus.PAID.value, "archived_at": None})
    
    # Calculate total outstanding
    pipeline = [
        {"$match": {"archived_at": None, "status": LoanStatus.OPEN.value}},
        {"$group": {"_id": None, "total": {"$sum": "$outstanding_balance"}}}
    ]
    result = await db.loans.aggregate(pipeline).to_list(1)
    total_outstanding = result[0]["total"] if result else 0
    
    # Fraud counts
    loans = await db.loans.find({"archived_at": None}, {"_id": 0}).to_list(10000)
    quick_close_count = 0
    duplicate_customers = set()
    customer_loan_count = {}
    
    for loan in loans:
        cid = loan["customer_id"]
        customer_loan_count[cid] = customer_loan_count.get(cid, 0) + 1
        
        if loan["status"] == LoanStatus.PAID.value:
            payments = await db.payments.find({"loan_id": loan["id"]}, {"_id": 0}).to_list(100)
            created_date = loan["created_at"][:10]
            last_payment = max([p["paid_at"][:10] for p in payments if p.get("paid_at")] or [None])
            if last_payment and created_date == last_payment:
                quick_close_count += 1
    
    for cid, count in customer_loan_count.items():
        if count > 1:
            duplicate_customers.add(cid)
    
    return {
        "total_customers": total_customers,
        "total_loans": total_loans,
        "open_loans": open_loans,
        "paid_loans": paid_loans,
        "total_outstanding": round(total_outstanding, 2),
        "quick_close_alerts": quick_close_count,
        "duplicate_customer_alerts": len(duplicate_customers)
    }

# ==================== DATABASE BACKUP ====================
class BackupRequest(BaseModel):
    backup_path: Optional[str] = None  # If not provided, uses configured path

class BackupResponse(BaseModel):
    success: bool
    message: str
    filename: Optional[str] = None
    filepath: Optional[str] = None
    backup_size: Optional[str] = None
    records_count: Optional[dict] = None

@api_router.get("/backup/status")
async def get_backup_status(user: dict = Depends(require_role(UserRole.ADMIN))):
    """Get backup configuration and last backup info"""
    backup_settings = await db.settings.find_one({"key": "backup_config"})
    last_backup = await db.settings.find_one({"key": "last_backup"})
    
    return {
        "backup_folder_path": backup_settings.get("value", {}).get("folder_path", "") if backup_settings else "",
        "auto_backup_enabled": backup_settings.get("value", {}).get("auto_backup", False) if backup_settings else False,
        "last_backup": last_backup.get("value", {}) if last_backup else None
    }

@api_router.put("/backup/config")
async def update_backup_config(
    folder_path: str = "",
    auto_backup: bool = False,
    user: dict = Depends(require_role(UserRole.ADMIN))
):
    """Update backup configuration"""
    config = {
        "folder_path": folder_path,
        "auto_backup": auto_backup
    }
    
    await db.settings.update_one(
        {"key": "backup_config"},
        {"$set": {"key": "backup_config", "value": config, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    await create_audit_log("settings", "backup_config", "update", user["id"], user["full_name"], after=config)
    
    return {"message": "Backup configuration updated successfully"}

@api_router.post("/backup/create", response_model=BackupResponse)
async def create_backup(
    data: BackupRequest = None,
    user: dict = Depends(require_role(UserRole.ADMIN))
):
    """Create a full database backup as JSON"""
    import os
    
    # Get backup path
    backup_path = None
    if data and data.backup_path:
        backup_path = data.backup_path
    else:
        backup_settings = await db.settings.find_one({"key": "backup_config"})
        if backup_settings and backup_settings.get("value", {}).get("folder_path"):
            backup_path = backup_settings["value"]["folder_path"]
    
    if not backup_path:
        raise HTTPException(status_code=400, detail="No backup path configured. Set path in Admin Panel → Backup Settings")
    
    # Verify path exists (or can be created)
    try:
        os.makedirs(backup_path, exist_ok=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot access backup folder: {str(e)}")
    
    # Collect all data
    try:
        backup_data = {
            "backup_info": {
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user["full_name"],
                "created_by_id": user["id"],
                "app_version": "1.0.0"
            },
            "users": [],
            "customers": [],
            "loans": [],
            "payments": [],
            "audit_logs": [],
            "settings": []
        }
        
        # Export users (without password hashes for security)
        users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(10000)
        backup_data["users"] = users
        
        # Export customers
        customers = await db.customers.find({}, {"_id": 0}).to_list(100000)
        backup_data["customers"] = customers
        
        # Export loans
        loans = await db.loans.find({}, {"_id": 0}).to_list(100000)
        backup_data["loans"] = loans
        
        # Export payments
        payments = await db.payments.find({}, {"_id": 0}).to_list(500000)
        backup_data["payments"] = payments
        
        # Export audit logs
        audit_logs = await db.audit_logs.find({}, {"_id": 0}).to_list(100000)
        backup_data["audit_logs"] = audit_logs
        
        # Export settings (except master password)
        settings = await db.settings.find({"key": {"$ne": MASTER_PASSWORD_HASH_KEY}}, {"_id": 0}).to_list(100)
        backup_data["settings"] = settings
        
        # Generate filename with timestamp and branch
        branch_settings = await db.settings.find_one({"key": "branch_name"})
        branch_name = branch_settings.get("value", "Main") if branch_settings else "Main"
        branch_name = branch_name.replace(" ", "_").replace("/", "_")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"EasyMoney_Backup_{branch_name}_{timestamp}.json"
        filepath = os.path.join(backup_path, filename)
        
        # Write backup file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, indent=2, default=str)
        
        # Calculate file size
        file_size = os.path.getsize(filepath)
        if file_size < 1024:
            size_str = f"{file_size} bytes"
        elif file_size < 1024 * 1024:
            size_str = f"{file_size / 1024:.1f} KB"
        else:
            size_str = f"{file_size / (1024 * 1024):.1f} MB"
        
        # Record counts
        records_count = {
            "users": len(backup_data["users"]),
            "customers": len(backup_data["customers"]),
            "loans": len(backup_data["loans"]),
            "payments": len(backup_data["payments"]),
            "audit_logs": len(backup_data["audit_logs"])
        }
        
        # Store last backup info
        last_backup_info = {
            "filename": filename,
            "filepath": filepath,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["full_name"],
            "size": size_str,
            "records": records_count
        }
        
        await db.settings.update_one(
            {"key": "last_backup"},
            {"$set": {"key": "last_backup", "value": last_backup_info, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        
        await create_audit_log("backup", "database", "create", user["id"], user["full_name"], 
                              after={"filename": filename, "records": records_count})
        
        return BackupResponse(
            success=True,
            message="Backup created successfully",
            filename=filename,
            filepath=filepath,
            backup_size=size_str,
            records_count=records_count
        )
        
    except Exception as e:
        logging.error(f"Backup failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Backup failed: {str(e)}")

@api_router.post("/backup/restore")
async def restore_backup(
    filepath: str,
    user: dict = Depends(require_role(UserRole.ADMIN))
):
    """Restore database from a backup file (CAUTION: This will overwrite current data)"""
    import os
    
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Backup file not found")
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            backup_data = json.load(f)
        
        # Validate backup structure
        required_keys = ["backup_info", "users", "customers", "loans", "payments"]
        for key in required_keys:
            if key not in backup_data:
                raise HTTPException(status_code=400, detail=f"Invalid backup file: missing {key}")
        
        # Create audit log before restore
        await create_audit_log("backup", "database", "restore_started", user["id"], user["full_name"],
                              after={"source_file": filepath, "backup_date": backup_data["backup_info"]["created_at"]})
        
        # Clear and restore customers
        await db.customers.delete_many({})
        if backup_data["customers"]:
            await db.customers.insert_many(backup_data["customers"])
        
        # Clear and restore loans
        await db.loans.delete_many({})
        if backup_data["loans"]:
            await db.loans.insert_many(backup_data["loans"])
        
        # Clear and restore payments
        await db.payments.delete_many({})
        if backup_data["payments"]:
            await db.payments.insert_many(backup_data["payments"])
        
        # Note: Users and audit logs are NOT restored for security
        # Admin must recreate users if needed
        
        await create_audit_log("backup", "database", "restore_completed", user["id"], user["full_name"],
                              after={
                                  "customers_restored": len(backup_data["customers"]),
                                  "loans_restored": len(backup_data["loans"]),
                                  "payments_restored": len(backup_data["payments"])
                              })
        
        return {
            "success": True,
            "message": "Database restored successfully",
            "restored": {
                "customers": len(backup_data["customers"]),
                "loans": len(backup_data["loans"]),
                "payments": len(backup_data["payments"])
            },
            "note": "Users were NOT restored for security. Please recreate user accounts if needed."
        }
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid backup file format")
    except Exception as e:
        logging.error(f"Restore failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")

# ==================== ROOT ====================
@api_router.get("/")
async def root():
    return {"message": "EasyMoneyLoans Desktop API", "version": "1.0.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
